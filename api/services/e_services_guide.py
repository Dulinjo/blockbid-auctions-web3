from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from api.core.processor import get_runtime_data_dir
from api.services.config import feature_enabled


KB_DIR = get_runtime_data_dir() / "admin-kb"
KB_FILE = KB_DIR / "e_services_kb.json"


@dataclass(slots=True)
class EServicesSearchResult:
    services: list[dict[str, Any]]
    contacts: list[dict[str, Any]]
    envelope_clues: list[dict[str, Any]]
    topk_metrics: dict[str, int]
    admin_kb_version: str
    used_fallback: bool
    fallbacks_used: list[str]


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _split_semicolon(value: Any) -> list[str]:
    text = _normalize_text(value)
    if not text:
        return []
    return [item.strip() for item in text.split(";") if item.strip()]


def _to_bool(value: Any, default: bool = False) -> bool:
    text = _normalize_text(value).lower()
    if text in {"1", "true", "yes", "on"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return default


def _build_search_key(parts: list[str]) -> str:
    joined = " ".join(part for part in parts if part).strip().lower()
    return re.sub(r"\s+", " ", joined)


def _sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [_normalize_text(cell) for cell in values[0]]
    rows: list[dict[str, Any]] = []
    for row in values[1:]:
        payload: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            payload[header] = row[idx] if idx < len(row) else None
        rows.append(payload)
    return rows


class EServicesGuideService:
    def __init__(self) -> None:
        self.enabled = feature_enabled("ENABLE_E_SERVICES_GUIDE", True)
        self.enable_admin_kb = feature_enabled("ENABLE_ADMIN_SERVICE_KB", True)
        self.enable_envelope_clues = feature_enabled("ENABLE_ENVELOPE_CLUES", True)
        self.include_drafts = feature_enabled("ENABLE_DRAFT_ESERVICES_IN_DEV", True)
        self._kb_cache: dict[str, Any] | None = None

    def import_eservices_knowledge_base_from_xlsx(self, file_path: str) -> dict[str, Any]:
        if not self.enable_admin_kb:
            return {"imported": False, "reason": "ENABLE_ADMIN_SERVICE_KB=false"}
        path = Path(file_path)
        if not path.exists():
            return {"imported": False, "reason": "xlsx file missing"}

        workbook = load_workbook(path, data_only=True)
        service_rows = _sheet_rows(workbook, "Services")
        contact_rows = _sheet_rows(workbook, "Contacts")
        envelope_rows = _sheet_rows(workbook, "Envelope_Clues")
        list_rows = _sheet_rows(workbook, "Lists")
        update_rows = _sheet_rows(workbook, "How_to_Update")

        services: list[dict[str, Any]] = []
        contacts: list[dict[str, Any]] = []
        envelope_clues: list[dict[str, Any]] = []
        incomplete_rows = 0

        for row in service_rows:
            service_id = _normalize_text(row.get("service_id") or row.get("serviceId"))
            service_name = _normalize_text(row.get("service_name") or row.get("serviceName"))
            status = (_normalize_text(row.get("status")) or "draft").lower()
            if not service_id and not service_name:
                continue
            if not service_id or not service_name:
                status = "draft"
                incomplete_rows += 1
            search_key = _normalize_text(row.get("search_key"))
            if not search_key:
                search_key = _build_search_key(
                    [
                        service_name,
                        _normalize_text(row.get("institution")),
                        _normalize_text(row.get("service_category")),
                        _normalize_text(row.get("life_event_category")),
                        _normalize_text(row.get("keywords_synonyms") or row.get("keywords/synonyms")),
                        _normalize_text(row.get("user_phrases_examples") or row.get("user phrases")),
                    ]
                )
            services.append(
                {
                    "serviceId": service_id or f"DRAFT-SRV-{len(services)+1}",
                    "serviceName": service_name,
                    "institution": _normalize_text(row.get("institution")),
                    "orgLevel": _normalize_text(row.get("org_level") or row.get("orgLevel")),
                    "serviceChannel": _normalize_text(row.get("service_channel") or row.get("serviceChannel")),
                    "serviceCategory": _normalize_text(
                        row.get("service_category") or row.get("serviceCategory")
                    ),
                    "lifeEventCategory": _normalize_text(
                        row.get("life_event_category") or row.get("lifeEventCategory")
                    ),
                    "routePriority": _normalize_text(row.get("route_priority") or row.get("routePriority"))
                    or "normal",
                    "status": status,
                    "keywordsSynonyms": _split_semicolon(
                        row.get("keywords_synonyms") or row.get("keywords/synonyms")
                    ),
                    "userPhrasesExamples": _split_semicolon(
                        row.get("user_phrases_examples") or row.get("user phrases")
                    ),
                    "readyToUseInstructionCopy": _normalize_text(
                        row.get("ready_to_use_instruction_copy")
                        or row.get("readyToUseInstructionCopy")
                    ),
                    "relatedContactIds": _split_semicolon(
                        row.get("related_contact_ids") or row.get("relatedContactIds")
                    ),
                    "relatedEnvelopeClueIds": _split_semicolon(
                        row.get("related_envelope_clue_ids") or row.get("relatedEnvelopeClueIds")
                    ),
                    "searchKey": search_key,
                    "draft": status != "active",
                }
            )

        for row in contact_rows:
            contact_id = _normalize_text(row.get("contact_id") or row.get("contactId"))
            org_name = _normalize_text(row.get("org_name") or row.get("orgName"))
            status = (_normalize_text(row.get("status")) or "draft").lower()
            if not contact_id and not org_name:
                continue
            if not contact_id:
                status = "draft"
                incomplete_rows += 1
            contacts.append(
                {
                    "contactId": contact_id or f"DRAFT-CNT-{len(contacts)+1}",
                    "institution": _normalize_text(row.get("institution")),
                    "orgName": org_name,
                    "orgLevel": _normalize_text(row.get("org_level") or row.get("orgLevel")),
                    "contactType": _normalize_text(row.get("contact_type") or row.get("contactType")),
                    "contactPreference": _normalize_text(
                        row.get("contact_preference") or row.get("contactPreference")
                    ),
                    "phone": _normalize_text(row.get("phone")),
                    "email": _normalize_text(row.get("email")),
                    "portalUrl": _normalize_text(row.get("portal_url") or row.get("portalUrl")),
                    "contactPurpose": _normalize_text(
                        row.get("contact_purpose") or row.get("contactPurpose")
                    ),
                    "scriptForAgentCopy": _normalize_text(
                        row.get("script_for_agent_copy") or row.get("scriptForAgentCopy")
                    ),
                    "fallbackIfUnreachable": _normalize_text(
                        row.get("fallback_if_unreachable") or row.get("fallbackIfUnreachable")
                    ),
                    "relatedServiceIds": _split_semicolon(
                        row.get("related_service_ids") or row.get("relatedServiceIds")
                    ),
                    "relatedEnvelopeClueIds": _split_semicolon(
                        row.get("related_envelope_clue_ids") or row.get("relatedEnvelopeClueIds")
                    ),
                    "status": status,
                    "draft": status != "active",
                }
            )

        for row in envelope_rows:
            clue_id = _normalize_text(row.get("envelope_clue_id") or row.get("envelopeClueId"))
            status = (_normalize_text(row.get("status")) or "draft").lower()
            if not clue_id and not _normalize_text(row.get("document_type")):
                continue
            if not clue_id:
                status = "draft"
                incomplete_rows += 1
            envelope_clues.append(
                {
                    "envelopeClueId": clue_id or f"DRAFT-ENV-{len(envelope_clues)+1}",
                    "documentType": _normalize_text(row.get("document_type") or row.get("documentType")),
                    "issuingBodyName": _normalize_text(
                        row.get("issuing_body_name") or row.get("issuingBodyName")
                    ),
                    "issuingBodyAbbrev": _normalize_text(
                        row.get("issuing_body_abbrev") or row.get("issuingBodyAbbrev")
                    ),
                    "visibleLabelForCaseNumber": _normalize_text(
                        row.get("visible_label_for_case_number")
                        or row.get("visibleLabelForCaseNumber")
                    ),
                    "upisnikCode": _normalize_text(row.get("upisnik_code") or row.get("upisnikCode")),
                    "caseNumberPatternExample": _normalize_text(
                        row.get("case_number_pattern_example")
                        or row.get("caseNumberPatternExample")
                    ),
                    "inferredCategory": _normalize_text(
                        row.get("inferred_category") or row.get("inferredCategory")
                    ),
                    "mappedServiceId": _normalize_text(
                        row.get("mapped_service_id") or row.get("mappedServiceId")
                    ),
                    "mappedContactIds": _split_semicolon(
                        row.get("mapped_contact_ids") or row.get("mappedContactIds")
                    ),
                    "readyToUseInstructionCopy": _normalize_text(
                        row.get("ready_to_use_instruction_copy")
                        or row.get("readyToUseInstructionCopy")
                    ),
                    "whatAgentShouldAskNext": _normalize_text(
                        row.get("what_agent_should_ask_next") or row.get("whatAgentShouldAskNext")
                    ),
                    "confidenceRule": _normalize_text(
                        row.get("confidence_rule") or row.get("confidenceRule")
                    ),
                    "status": status,
                    "draft": status != "active",
                }
            )

        kb = {
            "adminKnowledgeBaseVersion": f"kb-{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}",
            "importedAt": datetime.now(UTC).isoformat(),
            "sourceFile": str(path),
            "services": services,
            "contacts": contacts,
            "envelopeClues": envelope_clues,
            "lists": list_rows,
            "howToUpdate": update_rows,
            "incompleteRows": incomplete_rows,
        }
        KB_DIR.mkdir(parents=True, exist_ok=True)
        KB_FILE.write_text(json.dumps(kb, ensure_ascii=False, indent=2), encoding="utf-8")
        self._kb_cache = kb
        return {
            "imported": True,
            "services": len(services),
            "contacts": len(contacts),
            "envelopeClues": len(envelope_clues),
            "adminKnowledgeBaseVersion": kb["adminKnowledgeBaseVersion"],
            "incompleteRows": incomplete_rows,
        }

    def _default_seed_kb(self) -> dict[str, Any]:
        services = [
            {
                "serviceId": "SRV-001",
                "serviceName": "Ostavinski postupak — informacije i pokretanje",
                "institution": "Primer: Osnovni sud / Javni beležnik",
                "orgLevel": "local",
                "serviceChannel": "info + in_person",
                "serviceCategory": "inheritance",
                "lifeEventCategory": "death_in_family",
                "routePriority": "normal",
                "status": "draft",
                "keywordsSynonyms": ["ostavina", "nasledstvo", "smrt", "ostavinski postupak", "naslednik"],
                "userPhrasesExamples": [
                    "umro mi je otac",
                    "preminula mi je majka",
                    "nasledstvo",
                    "ostavina",
                    "šta radim posle smrti u porodici",
                ],
                "readyToUseInstructionCopy": (
                    "Na osnovu vašeg opisa, najrelevantnije su informacije o ostavinskom postupku. "
                    "Pripremite smrtovnicu, osnovne podatke o naslednicima i podatke o imovini ako ih imate. "
                    "Sledeći korak je da proverite nadležnost i način pokretanja postupka."
                ),
                "relatedContactIds": [],
                "relatedEnvelopeClueIds": [],
                "searchKey": "ostavinski postupak nasledstvo smrt sud beleznik",
                "draft": True,
            },
            {
                "serviceId": "SRV-002",
                "serviceName": "Prijava nasilja / hitno usmeravanje",
                "institution": "Primer: Policija / Centar za socijalni rad",
                "orgLevel": "local",
                "serviceChannel": "phone + in_person",
                "serviceCategory": "safety",
                "lifeEventCategory": "violence_or_threat",
                "routePriority": "critical",
                "status": "draft",
                "keywordsSynonyms": ["nasilje", "pretnja", "hitno", "policija", "csr"],
                "userPhrasesExamples": [
                    "bojim se",
                    "partner me maltretira",
                    "trpim nasilje",
                    "hitno",
                    "pretnje",
                ],
                "readyToUseInstructionCopy": (
                    "Ako postoji neposredna opasnost, odmah se obratite hitnim službama ili policiji. "
                    "Ako nije hitno, mogu vas uputiti na odgovarajuću službu za dalju podršku i informacije."
                ),
                "relatedContactIds": [],
                "relatedEnvelopeClueIds": [],
                "searchKey": "nasilje pretnja policija hitno csr bezbednost",
                "draft": True,
            },
            {
                "serviceId": "SRV-003",
                "serviceName": "Provera statusa predmeta po broju predmeta (upisniku)",
                "institution": "Primer: Sud / organ uprave",
                "orgLevel": "national/local",
                "serviceChannel": "info",
                "serviceCategory": "case_status",
                "lifeEventCategory": "received_letter_or_case_number",
                "routePriority": "high",
                "status": "draft",
                "keywordsSynonyms": [
                    "broj predmeta",
                    "upisnik",
                    "status predmeta",
                    "pisarnica",
                    "predmet",
                    "rešenje",
                    "poziv",
                ],
                "userPhrasesExamples": [
                    "stigla mi je koverta",
                    "imam broj predmeta",
                    "kako da proverim status",
                    "dokle je stiglo",
                    "šta znači p 123/2024",
                ],
                "readyToUseInstructionCopy": (
                    "Pošaljite ili prepišite naziv organa i broj predmeta tačno kako piše na dopisu. "
                    "Na osnovu toga mogu da vas usmerim gde se status proverava i koji kontakt je najrelevantniji."
                ),
                "relatedContactIds": ["CNT-001"],
                "relatedEnvelopeClueIds": ["ENV-001", "ENV-002"],
                "searchKey": "broj predmeta upisnik status pisarnica koverta poziv resenje",
                "draft": True,
            },
            {
                "serviceId": "SRV-004",
                "serviceName": "Tok predmeta javnog izvršitelja",
                "institution": "Javni izvršitelj",
                "orgLevel": "",
                "serviceChannel": "info",
                "serviceCategory": "",
                "lifeEventCategory": "",
                "routePriority": "normal",
                "status": "draft",
                "keywordsSynonyms": ["izvršitelj", "izvrsitelj", "tok predmeta", "izvršenje"],
                "userPhrasesExamples": [],
                "readyToUseInstructionCopy": "",
                "relatedContactIds": [],
                "relatedEnvelopeClueIds": [],
                "searchKey": "javni izvrsitelj tok predmeta izvrsenje",
                "draft": True,
            },
            {
                "serviceId": "SRV-005",
                "serviceName": "Elektronska oglasna tabla suda",
                "institution": "Javni izvršitelj, sud",
                "orgLevel": "",
                "serviceChannel": "e-service",
                "serviceCategory": "",
                "lifeEventCategory": "",
                "routePriority": "normal",
                "status": "draft",
                "keywordsSynonyms": ["oglasna tabla", "sud", "objava"],
                "userPhrasesExamples": [],
                "readyToUseInstructionCopy": "",
                "relatedContactIds": [],
                "relatedEnvelopeClueIds": [],
                "searchKey": "elektronska oglasna tabla sud objava",
                "draft": True,
            },
            {
                "serviceId": "SRV-006",
                "serviceName": "Elektronska oglasna tabla suda",
                "institution": "",
                "orgLevel": "",
                "serviceChannel": "e-service",
                "serviceCategory": "",
                "lifeEventCategory": "",
                "routePriority": "normal",
                "status": "draft",
                "keywordsSynonyms": ["oglasna tabla", "sud"],
                "userPhrasesExamples": [],
                "readyToUseInstructionCopy": "",
                "relatedContactIds": [],
                "relatedEnvelopeClueIds": [],
                "searchKey": "elektronska oglasna tabla sud",
                "draft": True,
            },
            {
                "serviceId": "SRV-007",
                "serviceName": "Registar neplaćenih novčanih kazni i drugih novčanih iznosa",
                "institution": "",
                "orgLevel": "",
                "serviceChannel": "info",
                "serviceCategory": "",
                "lifeEventCategory": "",
                "routePriority": "normal",
                "status": "draft",
                "keywordsSynonyms": ["neplacene kazne", "kazne", "novcani iznosi", "registar"],
                "userPhrasesExamples": [],
                "readyToUseInstructionCopy": "",
                "relatedContactIds": [],
                "relatedEnvelopeClueIds": [],
                "searchKey": "registar neplacenih novcanih kazni",
                "draft": True,
            },
            {
                "serviceId": "SRV-008",
                "serviceName": "Zahtev za rehabilitaciono obeštećenje",
                "institution": "Ministarstvo pravde",
                "orgLevel": "national",
                "serviceChannel": "info + in_person",
                "serviceCategory": "",
                "lifeEventCategory": "",
                "routePriority": "normal",
                "status": "draft",
                "keywordsSynonyms": ["rehabilitaciono obeštećenje", "rehabilitacija", "obestecenje"],
                "userPhrasesExamples": [],
                "readyToUseInstructionCopy": (
                    "Komisija za rehabilitaciono obeštećenje razmatra zahteve i priloženu dokumentaciju i "
                    "donosi odgovarajuće odluke. Kada Komisija donese odluku kojom usvoji zahtev za "
                    "rehabilitaciono obeštećenje predlaže i visinu naknade. Nakon donošenja odluke o usvajanju "
                    "zahteva, Ministarstvo podnosiocu zahteva dostavlja Predlog sporazuma o rehabilitacionom "
                    "obeštećenju sa propratnim aktom (u kome je precizno navedeno šta od dokumentacije treba "
                    "dostaviti Ministarstvu, ako postoji saglasnost u vezi sporazuma) i ostavlja rok u kome je "
                    "potrebno da se podnosilac zahteva izjasni da li prihvata predloženi sporazum. Ukoliko "
                    "Komisija donese odluku kojom nije usvojila zahtev za rehabilitaciono obeštećenje, podnosiocu "
                    "zahteva se dostavlja obrazložen odgovor sa navedenim razlozima koji su bili opredeljujući "
                    "za donošenje odluke."
                ),
                "relatedContactIds": [],
                "relatedEnvelopeClueIds": [],
                "searchKey": "rehabilitaciono obestecenje ministarstvo pravde zahtev",
                "draft": True,
            },
            {
                "serviceId": "SRV-009",
                "serviceName": "Zahtev za naknadu štete neosnovano osuđenih i neosnovano lišenih slobode",
                "institution": "Ministarstvo pravde",
                "orgLevel": "national",
                "serviceChannel": "info",
                "serviceCategory": "",
                "lifeEventCategory": "",
                "routePriority": "normal",
                "status": "draft",
                "keywordsSynonyms": ["naknada štete", "neosnovano osudjeni", "lisenih slobode"],
                "userPhrasesExamples": [],
                "readyToUseInstructionCopy": "",
                "relatedContactIds": [],
                "relatedEnvelopeClueIds": [],
                "searchKey": "naknada stete neosnovano osudjeni lisenih slobode",
                "draft": True,
            },
            {
                "serviceId": "SRV-010",
                "serviceName": "Elektronska javna prodaja",
                "institution": "Ministarstvo pravde",
                "orgLevel": "national",
                "serviceChannel": "e-service",
                "serviceCategory": "",
                "lifeEventCategory": "",
                "routePriority": "normal",
                "status": "draft",
                "keywordsSynonyms": ["javna prodaja", "elektronska prodaja", "licitacija"],
                "userPhrasesExamples": [],
                "readyToUseInstructionCopy": "",
                "relatedContactIds": [],
                "relatedEnvelopeClueIds": [],
                "searchKey": "elektronska javna prodaja licitacija",
                "draft": True,
            },
            {
                "serviceId": "SRV-011",
                "serviceName": "Legalizacija dokumenata",
                "institution": "Ministarstvo pravde",
                "orgLevel": "national",
                "serviceChannel": "info + in_person",
                "serviceCategory": "",
                "lifeEventCategory": "",
                "routePriority": "normal",
                "status": "draft",
                "keywordsSynonyms": ["legalizacija", "apostil", "dokumenti"],
                "userPhrasesExamples": [],
                "readyToUseInstructionCopy": "",
                "relatedContactIds": [],
                "relatedEnvelopeClueIds": [],
                "searchKey": "legalizacija dokumenata apostil",
                "draft": True,
            },
            {
                "serviceId": "SRV-012",
                "serviceName": "Tok predmeta prekršajnih sudova",
                "institution": "Ministarstvo pravde",
                "orgLevel": "national",
                "serviceChannel": "info",
                "serviceCategory": "",
                "lifeEventCategory": "",
                "routePriority": "normal",
                "status": "draft",
                "keywordsSynonyms": ["prekrsajni sud", "tok predmeta", "status"],
                "userPhrasesExamples": [],
                "readyToUseInstructionCopy": "",
                "relatedContactIds": [],
                "relatedEnvelopeClueIds": [],
                "searchKey": "tok predmeta prekrsajnih sudova",
                "draft": True,
            },
        ]
        contacts = [
            {
                "contactId": "CNT-001",
                "institution": "Primer institucija",
                "orgName": "Primer: Pisarnica organa",
                "orgLevel": "local",
                "contactType": "phone",
                "contactPurpose": "status_check_or_registry",
                "status": "draft",
                "relatedServiceIds": ["SRV-003"],
                "notesForAgent": "Kada korisnik ima broj predmeta i traži proveru statusa — uputi na pisarnicu.",
                "scriptForAgentCopy": (
                    "Za proveru statusa pripremite naziv organa i broj predmeta tačno kako piše na dopisu. "
                    "Ako imate samo deo oznake, i to može pomoći."
                ),
                "fallbackIfUnreachable": (
                    "Ako se niko ne javlja ili kontakt nije ažuran, proveriti website, centralu ili povezani "
                    "kontakt iz baze."
                ),
                "relatedEnvelopeClueIds": ["ENV-001", "ENV-002"],
                "contactPreference": "phone_first",
                "draft": True,
            },
            {
                "contactId": "CNT-002",
                "institution": "Primer institucija",
                "orgName": "Primer: eUprava podrška",
                "orgLevel": "national",
                "contactType": "email",
                "email": "podrska@example.rs",
                "contactPurpose": "technical_support",
                "status": "draft",
                "notesForAgent": (
                    "Kada korisnik ne može da pristupi e-servisu ili ima tehnički problem sa nalogom."
                ),
                "scriptForAgentCopy": (
                    "Ovo izgleda kao tehnički problem sa pristupom servisu. Korisnika uputiti na podršku i "
                    "zamoliti da pripremi opis greške, vreme pokušaja i eventualni screenshot."
                ),
                "fallbackIfUnreachable": (
                    "Ako email ne radi, ponuditi website ili drugi kanal podrške iz baze."
                ),
                "relatedServiceIds": [],
                "relatedEnvelopeClueIds": [],
                "contactPreference": "email_first",
                "draft": True,
            },
        ]
        envelope_clues = [
            {
                "envelopeClueId": "ENV-001",
                "documentType": "court_letter",
                "issuingBodyName": "Primer: Osnovni sud u ___",
                "issuingBodyAbbrev": "OSU",
                "visibleLabelForCaseNumber": "Broj predmeta / Posl. br.",
                "upisnikCode": "P",
                "caseNumberPatternExample": "P 123/2024",
                "inferredCategory": "status_of_case",
                "mappedServiceId": "SRV-003",
                "mappedContactIds": ["CNT-001"],
                "status": "draft",
                "whatUserCanNotice": (
                    "Na koverti/dopisu piše naziv suda i 'Broj predmeta' sa oznakom poput 'P 123/2024'."
                ),
                "whatAgentShouldAskNext": (
                    "Traži tačan naziv suda, pun broj predmeta, godinu i da li korisnik ima rešenje, poziv ili presudu."
                ),
                "portalOrRegistryToCheck": "javna pretraga ako postoji; u suprotnom pisarnica / kontakt organa",
                "readyToUseInstructionCopy": (
                    "Na dopisu izgleda da je u pitanju sudski predmet. Zapišite naziv suda i broj predmeta tačno "
                    "kako piše, pa proverite status preko nadležnog portala ili pisarnice."
                ),
                "confidenceRule": (
                    "Ako postoje naziv suda + oznaka P + broj/godina, visoka verovatnoća da treba rutirati na "
                    "proveru statusa predmeta."
                ),
                "draft": True,
            },
            {
                "envelopeClueId": "ENV-002",
                "documentType": "administrative_letter",
                "issuingBodyName": "Primer: Organ uprave ___",
                "issuingBodyAbbrev": "ORG",
                "visibleLabelForCaseNumber": "Broj:",
                "upisnikCode": "U",
                "caseNumberPatternExample": "U-456/2025",
                "inferredCategory": "status_of_case",
                "mappedServiceId": "SRV-003",
                "mappedContactIds": ["CNT-001"],
                "status": "draft",
                "whatUserCanNotice": "Piše naziv organa i 'Broj:' pa oznaka predmeta.",
                "whatAgentShouldAskNext": (
                    "Traži naziv organa, broj predmeta, datum dopisa i vrstu akta (rešenje, zaključak, obaveštenje)."
                ),
                "portalOrRegistryToCheck": "zvanični portal organa; eUprava; pisarnica",
                "readyToUseInstructionCopy": (
                    "Na dopisu izgleda da je u pitanju upravni predmet. Potrebni su naziv organa i broj predmeta "
                    "da bi se proverio status ili naredni korak."
                ),
                "confidenceRule": (
                    "Ako postoje naziv organa + Broj: + U/oznaka upravnog predmeta, rutirati na status predmeta "
                    "ili kontakt organa."
                ),
                "draft": True,
            },
        ]
        lists = [
            {"key": "org_level", "values": ["local", "regional", "national", "national/local", "varies"]},
            {
                "key": "service_channel",
                "values": [
                    "e-service",
                    "email",
                    "phone",
                    "in_person",
                    "info",
                    "info + in_person",
                    "phone + in_person",
                ],
            },
            {"key": "needs_eID", "values": ["yes", "no", "maybe"]},
            {"key": "status", "values": ["draft", "active", "paused", "deprecated"]},
            {"key": "route_priority", "values": ["critical", "high", "normal", "low"]},
            {"key": "sensitivity_level", "values": ["urgent_safety", "high_impact", "general_info"]},
            {"key": "status_check_supported", "values": ["yes", "no", "limited"]},
            {"key": "contact_preference", "values": ["portal_first", "phone_first", "email_first"]},
            {
                "key": "contact_type_extended",
                "values": ["portal", "registry", "hotline", "pisarnica", "support", "information"],
            },
        ]
        how_to_update = [
            "Services: jedan red = jedan servis.",
            "Popuniti svakodnevni jezik korisnika, routing pitanja i gotove instrukcije.",
            "status=active koristiti samo za proverene redove.",
            "Agent u produkciji koristi samo active redove.",
            "Gotovi tekstovi u ready_to_use_instruction_copy i script_for_agent_copy treba da imaju 2–5 kratkih rečenica.",
            "Kada se promeni URL, kontakt, taksa, uslov ili portal, ažurirati last_verified.",
            "Ako informacija nije proverena, status ostaje draft.",
            "Ako servis privremeno ne radi, status = paused.",
            "Ako je servis ukinut ili zamenjen, status = deprecated.",
            "Ako korisnik šalje kovertu/dopis, koristiti Envelope_Clues.",
            "Agent pita samo podatke vidljive na dopisu/koverti i ne traži nepotrebne lične podatke.",
        ]
        return {
            "adminKnowledgeBaseVersion": "seed-v1",
            "importedAt": datetime.now(UTC).isoformat(),
            "sourceFile": "embedded-seed",
            "services": services,
            "contacts": contacts,
            "envelopeClues": envelope_clues,
            "lists": lists,
            "howToUpdate": how_to_update,
            "incompleteRows": 0,
        }

    def _load_kb(self) -> dict[str, Any]:
        if self._kb_cache is not None:
            return self._kb_cache
        if KB_FILE.exists():
            try:
                payload = json.loads(KB_FILE.read_text(encoding="utf-8"))
                if isinstance(payload, dict):
                    self._kb_cache = payload
                    return payload
            except (json.JSONDecodeError, OSError):
                pass
        self._kb_cache = self._default_seed_kb()
        return self._kb_cache

    def _service_visible(self, row: dict[str, Any]) -> bool:
        status = str(row.get("status", "draft")).lower()
        if status == "active":
            return True
        if status == "paused":
            return True
        if status == "deprecated":
            return False
        return self.include_drafts

    def _search_services(
        self,
        kb: dict[str, Any],
        user_question: str,
        legal_area: str,
        e_service_intent: str,
        top_k: int,
    ) -> list[dict[str, Any]]:
        query = f"{user_question} {legal_area} {e_service_intent}".lower()
        query_tokens = {token for token in re.findall(r"[a-zA-ZčćžšđČĆŽŠĐ0-9]{2,}", query)}
        candidates: list[tuple[float, dict[str, Any]]] = []
        for row in kb.get("services", []):
            if not isinstance(row, dict):
                continue
            if not self._service_visible(row):
                continue
            status = str(row.get("status", "draft")).lower()
            if status == "paused":
                pause_penalty = 0.08
            else:
                pause_penalty = 0.0
            blob = " ".join(
                [
                    str(row.get("serviceName", "")),
                    str(row.get("institution", "")),
                    str(row.get("serviceCategory", "")),
                    str(row.get("lifeEventCategory", "")),
                    str(row.get("routePriority", "")),
                    str(row.get("searchKey", "")),
                    " ".join(str(v) for v in row.get("keywordsSynonyms", [])),
                    " ".join(str(v) for v in row.get("userPhrasesExamples", [])),
                ]
            ).lower()
            overlap = sum(1 for token in query_tokens if token in blob)
            priority_bonus = (
                0.25
                if str(row.get("routePriority", "")).lower() == "critical"
                else 0.14
                if str(row.get("routePriority", "")).lower() == "high"
                else 0.0
            )
            score = min(1.0, (overlap / 20.0) + priority_bonus + (0.06 if status == "active" else 0.0))
            score = max(0.0, score - pause_penalty)
            rated = dict(row)
            rated["relevanceScore"] = round(score, 3)
            candidates.append((score, rated))
        candidates.sort(key=lambda item: item[0], reverse=True)
        return [item[1] for item in candidates[: max(top_k, 1)]]

    def _extract_envelope_clues(self, kb: dict[str, Any], user_question: str) -> list[dict[str, Any]]:
        if not self.enable_envelope_clues:
            return []
        question = user_question.lower()
        looks_like_case_number = bool(re.search(r"\b[apu]\s*[-]?\s*\d{1,6}\s*/\s*(19|20)\d{2}\b", question))
        mentions_envelope = any(
            term in question
            for term in ("koverta", "dopis", "poziv", "rešenje", "resenje", "posl. br", "broj predmeta")
        )
        if not (looks_like_case_number or mentions_envelope):
            return []
        rows: list[dict[str, Any]] = []
        for row in kb.get("envelopeClues", []):
            if not isinstance(row, dict):
                continue
            status = str(row.get("status", "draft")).lower()
            if status == "deprecated":
                continue
            if status != "active" and not self.include_drafts:
                continue
            text_blob = " ".join(
                [
                    str(row.get("documentType", "")),
                    str(row.get("visibleLabelForCaseNumber", "")),
                    str(row.get("upisnikCode", "")),
                    str(row.get("caseNumberPatternExample", "")),
                ]
            ).lower()
            bonus = 0.18 if str(row.get("upisnikCode", "")).lower() in question else 0.0
            overlap = sum(1 for token in re.findall(r"[a-z0-9čćžšđ]{2,}", question) if token in text_blob)
            score = min(1.0, (overlap / 16.0) + bonus)
            rated = dict(row)
            rated["relevanceScore"] = round(score, 3)
            rows.append(rated)
        rows.sort(key=lambda item: float(item.get("relevanceScore", 0.0)), reverse=True)
        return rows[:2]

    def _resolve_contacts(
        self,
        kb: dict[str, Any],
        services: list[dict[str, Any]],
        envelope_clues: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        wanted_ids: set[str] = set()
        for service in services:
            for item in service.get("relatedContactIds", []):
                wanted_ids.add(str(item))
        for clue in envelope_clues:
            for item in clue.get("mappedContactIds", []):
                wanted_ids.add(str(item))
        contacts: list[dict[str, Any]] = []
        for row in kb.get("contacts", []):
            if not isinstance(row, dict):
                continue
            contact_id = str(row.get("contactId", ""))
            status = str(row.get("status", "draft")).lower()
            if status == "deprecated":
                continue
            if status != "active" and not self.include_drafts:
                continue
            if wanted_ids and contact_id not in wanted_ids:
                continue
            contacts.append(dict(row))
        return contacts[:3]

    def search_e_services_guide(
        self,
        payload: dict[str, Any],
    ) -> EServicesSearchResult:
        if not self.enabled:
            return EServicesSearchResult(
                services=[],
                contacts=[],
                envelope_clues=[],
                topk_metrics={
                    "eServicesInitialResultsCount": 0,
                    "eServicesRerankedResultsCount": 0,
                    "eServicesDisplayedResultsCount": 0,
                },
                admin_kb_version="disabled",
                used_fallback=True,
                fallbacks_used=["e-services-guide-disabled"],
            )
        kb = self._load_kb()
        question = _normalize_text(payload.get("userQuestion"))
        legal_area = _normalize_text(payload.get("legalArea"))
        e_service_intent = _normalize_text(payload.get("eServiceIntent"))
        top_k = max(int(payload.get("topK") or 10), 1)

        initial = self._search_services(kb, question, legal_area, e_service_intent, top_k=top_k)
        reranked = initial[: min(top_k, 3)]
        displayed = reranked[:2]
        envelope_clues = self._extract_envelope_clues(kb, question)
        contacts = self._resolve_contacts(kb, displayed, envelope_clues)
        return EServicesSearchResult(
            services=displayed,
            contacts=contacts,
            envelope_clues=envelope_clues,
            topk_metrics={
                "eServicesInitialResultsCount": len(initial),
                "eServicesRerankedResultsCount": len(reranked),
                "eServicesDisplayedResultsCount": len(displayed),
            },
            admin_kb_version=str(kb.get("adminKnowledgeBaseVersion", "seed-v1")),
            used_fallback=False,
            fallbacks_used=[],
        )


e_services_guide_service = EServicesGuideService()


def import_eservices_knowledge_base_from_xlsx(file_path: str) -> dict[str, Any]:
    return e_services_guide_service.import_eservices_knowledge_base_from_xlsx(file_path)


def search_e_services_guide(payload: dict[str, Any]) -> dict[str, Any]:
    result = e_services_guide_service.search_e_services_guide(payload)
    return {
        "services": result.services,
        "contacts": result.contacts,
        "envelopeClues": result.envelope_clues,
        "topKMetrics": result.topk_metrics,
        "adminKnowledgeBaseVersion": result.admin_kb_version,
        "usedFallback": result.used_fallback,
        "fallbacksUsed": result.fallbacks_used,
    }
